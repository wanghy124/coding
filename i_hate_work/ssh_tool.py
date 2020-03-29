# 1. ssh客户端
# 2. 从表格读取host ip
# 3. 多线程
# 4. 异常处理
# 5. 结果保存到txt，如有异常记录到表格

import paramiko
import time
import os
from tqdm import tqdm
from openpyxl import load_workbook
from multiprocessing.pool import ThreadPool
from multiprocessing import freeze_support

def get_host_info(file, sheet_name):
    data = load_workbook(file)  # 读取xlsx文件
    table = data[sheet_name]  # 读取sheet数据
    excel_dict = {}
    hostname = 'hostname'
    ip = 'ip'
    for row in table.iter_rows(min_row=2):
        cell_location = 1
        for cell in row:
            if cell_location == 1:  # 读取第一列的用户名
                hostname = cell.value
                cell_location += 1
            elif cell_location == 2:  # 读取第二列的密码
                ip = cell.value
                cell_location += 1
        excel_dict[hostname] = ip  # 写入字典
    return excel_dict  # 返回字典


def ssh_tool(hostname, ip, username, password, cmd_list, verbose=True):
    try:
        ssh = paramiko.SSHClient()  # 创建SSH Client
        ssh.load_system_host_keys()  # 加载系统SSH密钥
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())  # 添加新的SSH密钥
        ssh.connect(ip, port=22, username=username, password=password, timeout=5, compress=True)  # SSH连接

        chan = ssh.invoke_shell()  # 激活交互式shell
        time.sleep(1)
        x = chan.recv(2048).decode()  # 接收回显信息

        now = time.strftime("%Y_%m_%d_%H_%M_%S", time.localtime(time.time()))
        fname = hostname + r"_" + now + r".txt"

        with open(fname, "w", newline='') as file:
            for cmd in cmd_list:  # 读取命令
                chan.send(cmd.encode())  # 执行命令，注意字串都需要编码为二进制字串
                chan.send(b'\n')  # 一定要注意输入回车
                time.sleep(2)  # 由于有些回显可能过长，所以可以考虑等待更长一些时间
                x = chan.recv(40960).decode()  # 读取回显，有些回想可能过长，请把接收缓存调大
                if verbose:
                    # print(x)  # 打印回显
                    file.write(x)

        chan.close()  # 退出交互式shell
        ssh.close()  # 退出ssh会话
    except Exception as e:
        fname = r"Error_" + hostname + r".txt"
        with open(fname, "w", newline='') as file:
            file.write(hostname + '\n')
            file.write(ip + '\n')
            file.write(str(e))


if __name__ == '__main__':
    # 使用Linux解释器 & WIN解释器
    cmd_list = ['terminal length 0', 'show ver', 'show vlan']
    host_info = get_host_info('host.xlsx', 'Sheet1')
    freeze_support()
    pool = ThreadPool(processes=150)
    with tqdm(total=len(host_info)) as pbar:
        for hostname, ip in host_info.items():
            # ssh_tool(hostname, ip, 'netstar', 'netstar01', cmd_list)
            pool.apply_async(ssh_tool, args=(hostname, ip, 'netstar', 'netstar01', cmd_list))
            pbar.update(1)
    pool.close()
    pool.join()
    print('Finish')
    os.system('pause')

