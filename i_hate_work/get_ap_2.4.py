import paramiko
import time
import os
from openpyxl import load_workbook
import encodings.idna
import re
import csv
from tqdm import tqdm

def get_host_info(file, sheet_name):
    data = load_workbook(file)  # 读取xlsx文件
    table = data[sheet_name]  # 读取sheet数据
    excel_list = []
    hostname = 'hostname'
    ip = 'ip'
    username = 'username'
    password = 'password'
    for row in table.iter_rows(min_row=2):
        cell_location = 1
        for cell in row:
            if cell_location == 1:  # 读取第一列
                hostname = cell.value
                cell_location += 1
            elif cell_location == 2:  # 读取第二列
                ip = cell.value
                cell_location += 1
            elif cell_location == 3:  # 读取第三列
                username = cell.value
                cell_location += 1
            elif cell_location == 4:  # 读取第四列
                password = cell.value
                cell_location += 1
        excel_list.append([hostname, ip, username, password])
    return excel_list

def ssh_tool(hostname, ip, username, password, verbose=True):
    try:
        ssh = paramiko.SSHClient()  # 创建SSH Client
        ssh.load_system_host_keys()  # 加载系统SSH密钥
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())  # 添加新的SSH密钥
        ssh.connect(ip, port=22, username=username, password=password, timeout=10, compress=True, allow_agent=False,
                    look_for_keys=False)  # SSH连接

        chan = ssh.invoke_shell()  # 激活交互式shell
        time.sleep(1)

# --------------------获取AP列表--------------------

        ap_list = []
        ap_flex = {}

        chan.send(username)
        chan.send(b'\n')
        chan.send(password)
        chan.send(b'\n')
        time.sleep(3)
        chan.send(b'config paging disable')
        chan.send(b'\n')
        time.sleep(1)
        chan.send(b'show ap summary')
        chan.send(b'\n')
        time.sleep(3)
        output = chan.recv(200000).decode()  # 读取回显，有些回想可能过长，请把接收缓存调大
        # if verbose:
        #     print(output)  # 打印回显

        ap_summary = str(output).split('\n')
        for i in ap_summary:
            ap = re.match('(CN[-A-Za-z0-9]*)\s+', i)
            # ap = re.findall('([-A-Za-z0-9]+)\s+.*AIR', i)
            if ap:
                ap_list.append(ap[0])


# --------------------获取AP配置--------------------
        with tqdm(total=len(ap_list)) as pbar:
            for ap in ap_list:
                chan.send(b'show ap config 802.11b ')
                chan.send(ap)
                chan.send(b'\n')
                time.sleep(0.5)
                output2 = chan.recv(40960).decode()
                # if verbose:
                #     print(output2)

                ap_config = (str(output2)).replace('\n', ' ')
                radio = re.findall('.*RADIO_TYPE_80211n-2.4\s+Administrative State \.+\s(ADMIN_DISABLED|ADMIN_ENABLED).*', ap_config)
                if 'AP does not have the 802.11bg radio' in ap_config:
                    ap_flex[ap] = 'ADMIN_DISABLED'
                elif radio:
                    ap_flex[ap] = radio[0]
                else:
                    ap_flex[ap] = 'none'
                pbar.update(1)

        chan.send(b'config paging enable')
        chan.send(b'\n')
        time.sleep(1)

        chan.close()  # 退出交互式shell
        ssh.close()  # 退出ssh会话

        now = time.strftime("%Y_%m_%d_%H_%M_%S", time.localtime(time.time()))
        fname = now + r'_' + hostname + r"_AP_2.4GHz_Status.csv"
        with open(fname, "w", newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["AP Name", "2.4GHz Status"])
            for key, value in ap_flex.items():
                writer.writerow([key, value])

    except Exception as e:
        fname = r"Error_" + hostname + r".txt"
        with open(fname, "w", newline='') as file:
            file.write(hostname + '\n')
            file.write(ip + '\n')
            file.write(str(e))


if __name__ == '__main__':
    host_info = get_host_info('wlc.xlsx', 'Sheet1')
    for wlc in host_info:
        ssh_tool(wlc[0], wlc[1], wlc[2], wlc[3])
    print('Finish')
    os.system('pause')