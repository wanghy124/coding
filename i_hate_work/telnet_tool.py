# 1. ssh客户端
# 2. 从表格读取host ip
# 3. 多线程
# 4. 异常处理
# 5. 结果保存到txt，如有异常记录到表格
# -*- coding: utf-8 -*-

import time
import os
from tqdm import tqdm
from openpyxl import load_workbook
import encodings.idna
import telnetlib

def get_host_info(file, sheet_name):
    data = load_workbook(file)  # 读取xlsx文件
    table = data[sheet_name]  # 读取sheet数据
    excel_dict = {}
    hostname = 'hostname'
    ip = 'ip'
    for row in table.iter_rows(min_row=2):
        cell_location = 1
        for cell in row:
            if cell_location == 1:  # 读取第一列的用户名
                hostname = cell.value
                cell_location += 1
            elif cell_location == 2:  # 读取第二列的密码
                ip = cell.value
                cell_location += 1
        excel_dict[hostname] = ip  # 写入字典
    return excel_dict  # 返回字典

def get_cmd_list(file, sheet_name):
    data = load_workbook(file)
    table = data[sheet_name]
    excel_list = []
    cmd = 'cmd'
    for row in table.iter_rows(min_row=2):
        for cell in row:
            cmd = cell.value
        excel_list.append(cmd)
    return excel_list


def telnet_tool_1(hostname, ip, cmd_list):
    # try:
        # 连接Telnet服务器
        tn = telnetlib.Telnet(ip, port=23, timeout=10)
        # 输入登录用户名密码
        tn.read_until(b'sername: ')
        tn.write(b'netstar\n')
        tn.read_until(b'assword: ')
        tn.write(b'netstar01\n')
        time.sleep(1)
        tn.write(b'enable\n')
        tn.write(b'netstar01\n')
        time.sleep(1)

        # 登录完毕后执行命令
        fname = ip + r".txt"

        with open(fname, "w", newline='') as file:
            for cmd in cmd_list:
                tn.write(cmd.encode('ascii')+b'\n')
                time.sleep(3)
                result = ''
                while result.find('#') == -1:
                    result = tn.read_very_eager().decode('ascii')
                print(result)
                file.write(result)


        # 执行完毕后，终止Telnet连接（或输入exit退出）
        tn.close()  # tn.write('exit\n')

    # except Exception as e:
    #     fname = r"Error_" + ip + r".txt"
    #     with open(fname, "w", newline='') as file:
    #         file.write(hostname + '\n')
    #         file.write(ip + '\n')
    #         file.write(str(e))



if __name__ == '__main__':
    # 使用Linux解释器 & WIN解释器
    # cmd_list = ['terminal length 0', 'show run', 'show vlan', 'show inv', 'show ver']
    host_info = get_host_info('host.xlsx', 'Sheet1')
    cmd_list = get_cmd_list('host.xlsx', 'Sheet2')
    # print(cmd_list)
    # account_info = get_account_info('host.xlsx', 'Sheet3')
    # freeze_support()
    # pool = ThreadPool(processes=150)
    with tqdm(total=len(host_info)) as pbar:
        for hostname, ip in host_info.items():
            # result = pool.apply_async(ssh_tool, args=(hostname, ip, account[0], account[1], cmd_list))
            telnet_tool_1(hostname, ip, cmd_list)
        pbar.update(1)
    # pool.close()
    # pool.join()
    print('Finish')
    os.system('pause')

