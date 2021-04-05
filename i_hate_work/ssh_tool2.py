# 1. ssh客户端
# 2. 从表格读取host ip
# 3. 多线程
# 4. 异常处理
# 5. 结果保存到txt，如有异常记录到表格

import paramiko
import time
import os
from tqdm import tqdm
from openpyxl import load_workbook
from multiprocessing.pool import ThreadPool
from multiprocessing import freeze_support
import encodings.idna
import telnetlib

def get_host_info(file, sheet_name):
    data = load_workbook(file)  # 读取xlsx文件
    table = data[sheet_name]  # 读取sheet数据
    excel_dict = {}
    hostname = 'hostname'
    ip = 'ip'
    for row in table.iter_rows(min_row=2):
        cell_location = 1
        for cell in row:
            if cell_location == 1:  # 读取第一列的用户名
                hostname = cell.value
                cell_location += 1
            elif cell_location == 2:  # 读取第二列的密码
                ip = cell.value
                cell_location += 1
        excel_dict[hostname] = ip  # 写入字典
    return excel_dict  # 返回字典

def get_cmd_list(file, sheet_name):
    data = load_workbook(file)
    table = data[sheet_name]
    excel_list = []
    cmd = 'cmd'
    for row in table.iter_rows(min_row=2):
        for cell in row:
            cmd = cell.value
        excel_list.append(cmd)
    return excel_list

def get_account_info(file, sheet_name):
    data = load_workbook(file)
    table = data[sheet_name]
    excel_list = []
    username = 'username'
    password = 'password'
    for row in table.iter_rows(min_row=2):
        cell_location = 1
        for cell in row:
            if cell_location == 1:
                username = cell.value
                cell_location += 1
            elif cell_location == 2:
                password = cell.value
                cell_location += 1
        excel_list.append([username, password])
    return excel_list

def ssh_tool_1(hostname, ip, cmd_list, verbose=True):
    try:
        ssh = paramiko.SSHClient()  # 创建SSH Client
        ssh.load_system_host_keys()  # 加载系统SSH密钥
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())  # 添加新的SSH密钥
        ssh.connect(ip, port=22, username='netstar', password='netstar01', timeout=10, compress=True, allow_agent=False,
                    look_for_keys=False)  # SSH连接

        chan = ssh.invoke_shell()  # 激活交互式shell
        time.sleep(1)
        # x = chan.recv(2048).decode()  # 接收回显信息

        # now = time.strftime("%Y_%m_%d_%H_%M_%S", time.localtime(time.time()))
        fname = ip + r".txt"

        with open(fname, "w", newline='') as file:
            chan.send(b'enable\n')
            chan.send(b'netstar01\n')
            time.sleep(1)
            chan.send(b'terminal length 0\n')
            for cmd in cmd_list:  # 读取命令
                chan.send(cmd.encode())  # 执行命令，注意字串都需要编码为二进制字串
                chan.send(b'\n')  # 一定要注意输入回车
                time.sleep(3)  # 由于有些回显可能过长，所以可以考虑等待更长一些时间
                x = chan.recv(200000).decode()  # 读取回显，有些回想可能过长，请把接收缓存调大
                if verbose:
                    print(x)  # 打印回显
                    file.write(x)

        chan.close()  # 退出交互式shell
        ssh.close()  # 退出ssh会话

    except Exception as e:
        if 'Authentication failed' in str(e):
            ssh_tool_2(hostname, ip, cmd_list)
        elif 'Unable to connect to port 22' in str(e):
            telnet_tool_1(hostname, ip, cmd_list)
        else:
            fname = r"Error_" + ip + r".txt"
            with open(fname, "w", newline='') as file:
                file.write(hostname + '\n')
                file.write(ip + '\n')
                file.write(str(e))


def ssh_tool_2(hostname, ip, cmd_list, verbose=True):
    try:
        ssh = paramiko.SSHClient()  # 创建SSH Client
        ssh.load_system_host_keys()  # 加载系统SSH密钥
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())  # 添加新的SSH密钥
        ssh.connect(ip, port=22, username='cn-admin', password='2018CHINAteam@06', timeout=10, compress=True, allow_agent=False,
                    look_for_keys=False)  # SSH连接

        chan = ssh.invoke_shell()  # 激活交互式shell
        time.sleep(1)

        fname = ip + r".txt"

        with open(fname, "w", newline='') as file:
            chan.send(b'enable\n')
            chan.send(b'2018CHINAteam@06\n')
            time.sleep(1)
            chan.send(b'terminal length 0\n')
            for cmd in cmd_list:  # 读取命令
                chan.send(cmd.encode())  # 执行命令，注意字串都需要编码为二进制字串
                chan.send(b'\n')  # 一定要注意输入回车
                time.sleep(3)  # 由于有些回显可能过长，所以可以考虑等待更长一些时间
                x = chan.recv(200000).decode()  # 读取回显，有些回想可能过长，请把接收缓存调大
                if verbose:
                    print(x)  # 打印回显
                    file.write(x)

        chan.close()  # 退出交互式shell
        ssh.close()  # 退出ssh会话

    except Exception as e:
        fname = r"Error_" + ip + r".txt"
        with open(fname, "w", newline='') as file:
            file.write(hostname + '\n')
            file.write(ip + '\n')
            file.write(str(e))


def telnet_tool_1(hostname, ip, cmd_list):
    try:
        # 连接Telnet服务器
        tn = telnetlib.Telnet(ip, port=23, timeout=10)
        # 输入登录用户名密码
        tn.read_until(b'sername: ')
        tn.write(b'netstar\n')
        tn.read_until(b'assword: ')
        tn.write(b'netstar01\n')
        time.sleep(1)
        tn.write(b'enable\n')
        tn.write(b'netstar01\n')
        time.sleep(1)
        tn.write(b'terminal length 0\n')

        # 登录完毕后执行命令
        fname = ip + r".txt"

        with open(fname, "w", newline='') as file:
            for cmd in cmd_list:
                tn.write(cmd.encode('ascii')+b'\n')
                time.sleep(3)
                result = ''
                while result.find('#') == -1:
                    result = tn.read_very_eager().decode('ascii')
                print(result)
                file.write(result)

        # 执行完毕后，终止Telnet连接（或输入exit退出）
        tn.close()  # tn.write('exit\n')

    except Exception as e:
        if 'Authentication failed' in str(e):
            telnet_tool_2(hostname, ip, cmd_list)
        else:
            fname = r"Error_" + ip + r".txt"
            with open(fname, "w", newline='') as file:
                file.write(hostname + '\n')
                file.write(ip + '\n')
                file.write(str(e))


def telnet_tool_2(hostname, ip, cmd_list):
    try:
        # 连接Telnet服务器
        tn = telnetlib.Telnet(ip, port=23, timeout=10)
        # 输入登录用户名密码
        tn.read_until(b'sername: ')
        tn.write(b'cn-admin\n')
        tn.read_until(b'assword: ')
        tn.write(b'2018CHINAteam@06\n')
        time.sleep(1)
        tn.write(b'enable\n')
        tn.write(b'2018CHINAteam@06\n')
        time.sleep(1)
        tn.write(b'terminal length 0\n')

        # 登录完毕后执行命令
        fname = ip + r".txt"

        with open(fname, "w", newline='') as file:
            for cmd in cmd_list:
                tn.write(cmd.encode('ascii')+b'\n')
                time.sleep(3)
                result = ''
                while result.find('#') == -1:
                    result = tn.read_very_eager().decode('ascii')
                print(result)
                file.write(result)

        # 执行完毕后，终止Telnet连接（或输入exit退出）
        tn.close()  # tn.write('exit\n')

    except Exception as e:
        fname = r"Error_" + ip + r".txt"
        with open(fname, "w", newline='') as file:
            file.write(hostname + '\n')
            file.write(ip + '\n')
            file.write(str(e))


if __name__ == '__main__':
    # 使用Linux解释器 & WIN解释器
    # cmd_list = ['terminal length 0', 'show run', 'show vlan', 'show inv', 'show ver']
    host_info = get_host_info('host.xlsx', 'Sheet1')
    cmd_list = get_cmd_list('host.xlsx', 'Sheet2')
    # print(cmd_list)
    # account_info = get_account_info('host.xlsx', 'Sheet3')
    # freeze_support()
    # pool = ThreadPool(processes=150)
    with tqdm(total=len(host_info)) as pbar:
        for hostname, ip in host_info.items():
            # result = pool.apply_async(ssh_tool, args=(hostname, ip, account[0], account[1], cmd_list))
            ssh_tool_1(hostname, ip, cmd_list)
            pbar.update(1)
    # pool.close()
    # pool.join()
    print('Finish')
    os.system('pause')

