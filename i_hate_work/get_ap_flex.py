import paramiko
import time
import os
from openpyxl import load_workbook
import encodings.idna
import re
import csv

def get_host_info(file, sheet_name):
    data = load_workbook(file)  # 读取xlsx文件
    table = data[sheet_name]  # 读取sheet数据
    excel_list = []
    hostname = 'hostname'
    ip = 'ip'
    username = 'username'
    password = 'password'
    for row in table.iter_rows(min_row=2):
        cell_location = 1
        for cell in row:
            if cell_location == 1:  # 读取第一列
                hostname = cell.value
                cell_location += 1
            elif cell_location == 2:  # 读取第二列
                ip = cell.value
                cell_location += 1
            elif cell_location == 3:  # 读取第三列
                username = cell.value
                cell_location += 1
            elif cell_location == 4:  # 读取第四列
                password = cell.value
                cell_location += 1
        excel_list.append([hostname, ip, username, password])
    return excel_list

def ssh_tool(hostname, ip, username, password, verbose=True):
    try:
        ssh = paramiko.SSHClient()  # 创建SSH Client
        ssh.load_system_host_keys()  # 加载系统SSH密钥
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())  # 添加新的SSH密钥
        ssh.connect(ip, port=22, username=username, password=password, timeout=5, compress=True, allow_agent=False,
                    look_for_keys=False)  # SSH连接

        chan = ssh.invoke_shell()  # 激活交互式shell
        time.sleep(1)

        now = time.strftime("%Y_%m_%d_%H_%M_%S", time.localtime(time.time()))
        fname = now + hostname + r".txt"

        with open(fname, "w", newline='') as file:
            chan.send(username)
            chan.send(b'\n')
            chan.send(password)
            chan.send(b'\n')
            time.sleep(3)
            chan.send(b'config paging disable')
            chan.send(b'\n')
            time.sleep(1)
            chan.send(b'show run-config')
            chan.send(b'\n')
            time.sleep(10)
            chan.send(b'config paging enable')
            chan.send(b'\n')
            time.sleep(1)
            x = chan.recv(10000000).decode()  # 读取回显，有些回想可能过长，请把接收缓存调大
            if verbose:
                print(x)  # 打印回显
                file.write(x)

        chan.close()  # 退出交互式shell
        ssh.close()  # 退出ssh会话
        return x

    except Exception as e:
        fname = r"Error_" + hostname + r".txt"
        with open(fname, "w", newline='') as file:
            file.write(hostname + '\n')
            file.write(ip + '\n')
            file.write(str(e))

def get_ap_config(wlc_config):
    ap_config_all = ''.join(re.findall('.*AP Config(.*)AP Airewave Director Configuration.*', wlc_config))
    ap_config = re.split('Cisco AP Identifier', ap_config_all)
    return ap_config

def get_ap_flex(ap_config):
    ap_flex = {}
    for ap in ap_config:
        ap_name = re.findall('.*Cisco AP Name\.+\s(\w+)\s.*', ap)
        native_id = re.findall('.*Native ID :\.+\s(\d+)\s.*', ap)
        if native_id:
            ap_flex[ap_name[0]] = native_id[0]
        elif ap_name:
            ap_flex[ap_name[0]] = 'none'
    return ap_flex

if __name__ == '__main__':
    with open('ap_flex.txt', 'r') as f:
        config = f.read().replace('\n', ' ')
    ap_config = get_ap_config(config)
    print(get_ap_flex(ap_config))


