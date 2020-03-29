#!/usr/bin/env python3
# -*- coding=utf-8 -*-
# 本脚由亁颐堂现任明教教主编写，用于乾颐盾Python课程！
# 教主QQ:605658506
# 亁颐堂官网www.qytang.com
# 教主技术进化论拓展你的技术新边疆
# https://ke.qq.com/course/271956?tuin=24199d8a

from openpyxl import Workbook
from openpyxl import load_workbook


def get_host_info(file, sheet_name):
    data = load_workbook(file)  # 读取xlsx文件
    table = data[sheet_name]  # 读取sheet数据
    excel_dict = {}
    hostname = 'hostname'
    ip = 'ip'
    for row in table.iter_rows(min_row=2):
        cell_location = 1
        for cell in row:
            if cell_location == 1:  # 读取第一列的用户名
                hostname = cell.value
                cell_location += 1
            elif cell_location == 2:  # 读取第二列的密码
                ip = cell.value
                cell_location += 1
        excel_dict[hostname] = ip  # 写入字典
    return excel_dict  # 返回字典


if __name__ == "__main__":
    print(get_host_info('host.xlsx', 'Sheet1'))
